import os
import time
from datetime import datetime
from io import BytesIO
import logging
import requests

from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_wtf.csrf import CSRFProtect
from flask_reuploads import UploadSet, configure_uploads, IMAGES

from sqlalchemy import func, desc
from werkzeug.utils import secure_filename

from openpyxl import Workbook

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader

from models import FailedLoginAttempt, UsernameChangeLog, PasswordChangeLog, UserProfile, UserRole, Candidate, db, User, Vote
from config import Config
from forms import ChangePasswordForm, EditUsernameForm, EditProfileForm, LoginForm, RegisterForm, RegisterCandidate

app = Flask(__name__)
app.config.from_object(Config)

csrf = CSRFProtect(app)
db.init_app(app)

app.config['UPLOADED_PHOTOS_DEST'] = 'uploads/photos'
photos = UploadSet('photos', IMAGES)
configure_uploads(app, photos)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = "Debes iniciar sesión para acceder a esta página."
login_manager.login_message_category = "warning"

@login_manager.user_loader
def load_user(user_id):
    """
    Load a user from the database by user ID for Flask-Login.

    Args:
        user_id: The ID of the user to load

    Returns:
        User: The User object if found, None otherwise
    """
    return db.session.get(User, int(user_id))

@app.route('/uploads/photos/<filename>')
def uploaded_file(filename):
    """
    Serve uploaded profile photos.

    Args:
        filename: The name of the file to serve

    Returns:
        File: The requested file from the uploads directory
    """
    return send_from_directory(app.config['UPLOADED_PHOTOS_DEST'], filename)

with app.app_context():
    db.create_all()

@app.route('/')
def home():
    """
    Render the home page (landing page).

    Returns:
        HTML: The rendered home.html template
    """
    return render_template('home.html')

@app.route('/register_candidate', methods=['GET', 'POST'])
@login_required
def register_candidate():
    """
    Register a new candidate for the election (Admin only).

    Allows admin users to register candidates with name, last name, proposal,
    and optional profile picture. The picture is saved to static/candidate_photos/.

    Returns:
        GET: Renders the candidate registration form
        POST: Processes form submission and redirects to login on success

    Access Control:
        Only users with ADMIN role can access this route
    """
    if current_user.role != UserRole.ADMIN:
        flash("No tienes permisos para registrar candidatos.", "danger")
        return redirect(url_for('index'))
    
    form = RegisterCandidate()
    if form.validate_on_submit():
        name = form.name.data
        last_name = form.last_name.data
        propuesta = form.propuesta.data
        
        profile_pic_filename = None
        if form.profile_picture.data:
            picture = form.profile_picture.data
            profile_pic_filename = secure_filename(f"{name}_{last_name}_{int(time.time())}.{picture.filename.rsplit('.', 1)[1].lower()}")
            
            upload_folder = os.path.join(app.root_path, 'static/candidate_photos')
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)
            
            picture.save(os.path.join(upload_folder, profile_pic_filename))
        
        new_candidate = Candidate(
            name=name,
            last_name=last_name,
            propuesta=propuesta,
            profile_picture=profile_pic_filename
        )
        
        db.session.add(new_candidate)
        db.session.commit()
        
        flash("Candidato registrado con éxito", "success")
        return redirect(url_for('login')) 
    
    return render_template('register_candidate.html', form=form)

@app.route('/register/<role>', methods=['GET', 'POST'])
@login_required
def register(role):
    """
    Register a new user with a specified role (Admin only).

    Args:
        role: The role to assign to the new user (docente, admin, or estudiante)

    Returns:
        GET: Renders the user registration form
        POST: Creates new user and their profile, redirects to login on success

    Access Control:
        Only users with ADMIN role can access this route

    Notes:
        Automatically creates a UserProfile entry for the new user
    """
    if current_user.role != UserRole.ADMIN:
        flash("No tienes permisos para crear usuarios.", "danger")
        return redirect(url_for('index'))
    
    form = RegisterForm()
    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data

        if User.query.filter_by(username=username).first():
            flash("El usuario ya existe", "danger")
            return redirect(url_for('register', role=role))

        if role not in ["docente", "admin", "estudiante"]:
            flash("Rol inválido", "danger")
            return redirect(url_for('home'))

        new_user = User(username=username, role=role)
        new_user.set_password(password)

        db.session.add(new_user)
        db.session.commit()

        user_profile = UserProfile(
            user_name=username,
            user_id=new_user.id,
        )
        db.session.add(user_profile)
        db.session.commit()

        flash("Usuario registrado con éxito", "success")
        return redirect(url_for('login'))

    return render_template('register.html', form=form, role=role)

MAX_ATTEMPTS = 3
BLOCK_TIME = 15
RECAPTCHA_SECRET_KEY = '6LfzLRsrAAAAALyqQGFcF0LFAHBPavE_lqE0yAhD'

@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    Handle user login with reCAPTCHA verification and failed attempt tracking.

    Implements security features:
    - reCAPTCHA verification
    - Failed login attempt tracking (max 3 attempts)
    - Automatic blocking for 15 seconds after max failed attempts
    - Password verification using Werkzeug

    Returns:
        GET: Renders the login form
        POST: Authenticates user and redirects to index on success

    Security:
        - Tracks failed attempts per username in FailedLoginAttempt table
        - Blocks further attempts for BLOCK_TIME seconds after MAX_ATTEMPTS failures
        - Resets counter on successful login
    """
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    form = LoginForm()

    if request.method == 'POST':
        recaptcha_response = request.form.get('g-recaptcha-response')
        payload = {
            'secret': RECAPTCHA_SECRET_KEY,
            'response': recaptcha_response
        }
        r = requests.post('https://www.google.com/recaptcha/api/siteverify', data=payload)
        result = r.json()

        if not result.get('success'):
            flash('reCAPTCHA falló. Inténtalo de nuevo.', 'danger')
            return redirect(url_for('login'))

        username = request.form['username']
        password = request.form['password']
        
        user = User.query.filter_by(username=username).first()
        failed_attempt = FailedLoginAttempt.query.filter_by(username=username).first()

        current_time = int(time.time())

        if not failed_attempt:
            failed_attempt = FailedLoginAttempt(username=username, attempts=0, last_attempt=0)
            db.session.add(failed_attempt)
            db.session.commit()

        time_since_last_attempt = current_time - failed_attempt.last_attempt
        if failed_attempt.attempts >= MAX_ATTEMPTS:
            if time_since_last_attempt < BLOCK_TIME:
                remaining_time = BLOCK_TIME - time_since_last_attempt
                flash(f"Demasiados intentos fallidos. Inténtalo en {remaining_time} segundos.", "danger")
                return redirect(url_for('login'))
            else:
                failed_attempt.attempts = 0
                failed_attempt.last_attempt = 0
                db.session.commit()

        if user and user.check_password(password):
            login_user(user)
            failed_attempt.attempts = 0
            failed_attempt.last_attempt = 0
            db.session.commit()
            flash("Inicio de sesión exitoso", "success")
            return redirect(url_for('index'))
        else:
            failed_attempt.attempts += 1
            failed_attempt.last_attempt = current_time
            db.session.commit()
            flash("Usuario o contraseña incorrectos", "danger")

    return render_template('login.html', form=form)

@app.route('/index')
@login_required
def index():
    return render_template('index.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash("Has cerrado sesión", "info")
    return redirect(url_for('login'))

@login_required
@app.route('/candidates')
def candidates_list():
    candidates = Candidate.query.all()
    return render_template('candidates_list.html', candidates=candidates)

@app.route('/view_profile', methods=['GET', 'POST'])
@login_required
def view_profile():
    return render_template('view_profile.html')

@app.route('/edit_profile', methods=['GET', 'POST'])
@login_required
def edit_profile():
    form = EditProfileForm()

    if request.method == 'GET':
        csrf_token = form.csrf_token._value()
        logging.info("Acceso a la pagina de edicion del perfil")
        logging.info(f'Token CSRF enviado del servidor al cliente: {csrf_token}')

    if not current_user.profile:
        profile = UserProfile(user_id=current_user.id)
        db.session.add(profile)
        db.session.commit()

    if form.validate_on_submit():
        email_exists = UserProfile.query.filter(
            UserProfile.email == form.email.data,
            UserProfile.user_id != current_user.id
        ).first()

        if email_exists:
            flash('Este correo ya está registrado por otro usuario.', 'danger')
            return redirect(url_for('edit_profile'))

        if form.profile_picture.data:
            filename = secure_filename(form.profile_picture.data.filename)
            allowed_extensions = {'jpg', 'jpeg', 'png', 'webp'}

            if filename.split('.')[-1].lower() not in allowed_extensions:
                flash('Formato de imagen no válido. Solo se permiten archivos JPG, JPEG, PNG y WEBP.', 'danger')
                return redirect(url_for('edit_profile'))

            upload_folder = app.config['UPLOADED_PHOTOS_DEST']
            os.makedirs(upload_folder, exist_ok=True)
            filepath = os.path.join(upload_folder, filename)
            form.profile_picture.data.save(filepath)
            current_user.profile.profile_picture = filename

        current_user.profile.email = form.email.data
        current_user.profile.name = form.name.data
        current_user.profile.last_name = form.last_name.data

        db.session.commit()
        flash('¡Perfil actualizado con éxito!', 'success')
        return redirect(url_for('index'))

    if current_user.profile:
        form.email.data = current_user.profile.email
        form.name.data = current_user.profile.name
        form.last_name.data = current_user.profile.last_name

    return render_template('edit_profile.html', form=form)

@app.route('/vote', methods=['GET', 'POST'])
@login_required
def vote():
    """
    Allow students to cast their vote for a candidate.

    Displays all candidates and allows the authenticated student to vote for one.
    Enforces one vote per student constraint.

    Returns:
        GET: Renders the voting page with candidate list
        POST: Records the vote and redirects to results page

    Access Control:
        Only users with ESTUDIANTE role can vote

    Security:
        - Prevents duplicate voting (one vote per user_id)
        - Vote is anonymous (stored with user_id but results are aggregated)
    """
    print(f"Current user role: {current_user.role}")
    print(f"Role type: {type(current_user.role)}")
    
    if hasattr(UserRole, 'ESTUDIANTE'):
        allowed_role = UserRole.ESTUDIANTE
    else:
        allowed_role = "estudiante"
    
    if current_user.role != allowed_role:
        if isinstance(current_user.role, str):
            if current_user.role.lower() != "estudiante":
                flash("Solo los estudiantes pueden votar", "danger")
                return redirect(url_for('index'))
        else:
            flash("Solo los estudiantes pueden votar", "danger")
            return redirect(url_for('index'))
    
    if Vote.query.filter_by(user_id=current_user.id).first():
        flash("Ya has emitido tu voto", "warning")
        return redirect(url_for('results'))
    
    candidates = Candidate.query.all()
    
    if request.method == 'POST':
        candidate_id = request.form.get('candidate_id')
        
        if not candidate_id:
            flash("Debes seleccionar un candidato", "danger")
            return redirect(url_for('vote'))
        
        try:
            new_vote = Vote(user_id=current_user.id, candidate_id=candidate_id)
            db.session.add(new_vote)
            db.session.commit()
            flash("¡Tu voto ha sido registrado exitosamente!", "success")
            return redirect(url_for('results'))
        except Exception as e:
            db.session.rollback()
            print(f"Error creating vote: {str(e)}")  
            flash("Ha ocurrido un error al registrar tu voto", "danger")
            return redirect(url_for('vote'))
    
    return render_template('vote.html', candidates=candidates)

@app.route('/results')
def results():
    """
    Display election results with vote counts and percentages.

    Aggregates votes for each candidate and calculates percentages.
    Results are ordered by vote count (highest first).

    Returns:
        HTML: Rendered results.html with:
            - List of candidates with vote counts
            - Percentage of total votes for each candidate
            - Total vote count
            - Whether current user has voted (if authenticated)

    Access Control:
        Public route (no login required)
    """
    candidates_with_votes = db.session.query(
        Candidate,
        func.count(Vote.id).label('vote_count')
    ).outerjoin(Vote).group_by(Candidate.id).order_by(desc('vote_count')).all()
    
    total_votes = Vote.query.count()
    
    results_data = []
    for candidate, vote_count in candidates_with_votes:
        percentage = (vote_count / total_votes * 100) if total_votes > 0 else 0
        results_data.append({
            'candidate': candidate,
            'votes': vote_count,
            'percentage': round(percentage, 1)
        })
    
    has_voted = False
    if current_user.is_authenticated:
        has_voted = Vote.query.filter_by(user_id=current_user.id).first() is not None
    
    return render_template('results.html', results=results_data, total_votes=total_votes, has_voted=has_voted)

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    form = ChangePasswordForm()

    if form.validate_on_submit():
        current_user.set_password(form.new_password.data)
        
        log = PasswordChangeLog(
            changed_by_user_id=current_user.id,
            affected_user_id=current_user.id
        )
        db.session.add(log)
        
        db.session.commit()
        flash("Contraseña actualizada exitosamente", "success")
        return redirect(url_for('index'))

    return render_template('change_password.html', form=form)

@app.route('/edit_user_password/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user_password(user_id):
    if current_user.role != UserRole.ADMIN:
        flash("No tienes permisos para editar contraseñas.", "danger")
        return redirect(url_for('index'))

    user = User.query.get_or_404(user_id)
    form = ChangePasswordForm()

    if form.validate_on_submit():
        user.set_password(form.new_password.data)
        
        log = PasswordChangeLog(
            changed_by_user_id=current_user.id,
            affected_user_id=user.id
        )
        db.session.add(log)
        
        db.session.commit()
        flash(f"Contraseña de {user.username} actualizada", "success")
        return redirect(url_for('index'))

    return render_template('edit_user_password.html', form=form, user=user)

@app.route('/manage_users')
@login_required
def manage_users():
    """
    Display user management dashboard (Admin only).

    Allows admins to view, search, and filter all users in the system.
    Supports search by username and filtering by role.

    Query Parameters:
        search: Optional username search term
        role: Optional role filter (admin, docente, estudiante)

    Returns:
        HTML: Rendered manage_users.html with user list

    Access Control:
        Only users with ADMIN role can access this route
    """
    if current_user.role != UserRole.ADMIN:
        flash("Acceso denegado. Solo el administrador puede ver esta sección.", "danger")
        return redirect(url_for('index'))

    search = request.args.get('search', '', type=str)
    role_filter = request.args.get('role', '', type=str)

    query = User.query.filter(User.id != current_user.id)

    if search:
        query = query.filter(
            (User.username.ilike(f"%{search}%"))
        )

    if role_filter:
        query = query.filter_by(role=UserRole(role_filter))

    users = query.all()
    return render_template('manage_users.html', users=users, search=search, role_filter=role_filter)


@app.route('/delete_user/<int:user_id>', methods=['POST']) 
@login_required
def delete_user(user_id):
    if current_user.role != UserRole.ADMIN:
        flash("Acceso denegado. Solo el administrador puede eliminar usuarios.", "danger")
        return redirect(url_for('index'))

    user = User.query.get_or_404(user_id)

    if user.id == current_user.id:
        flash("No puedes eliminar tu propia cuenta.", "warning")
        return redirect(url_for('manage_users'))

    db.session.delete(user)
    db.session.commit()
    flash(f"Usuario {user.username} eliminado exitosamente.", "success")
    return redirect(url_for('manage_users'))


@app.route('/edit_username', methods=['GET', 'POST'])
@login_required
def edit_username():
    form = EditUsernameForm()

    if form.validate_on_submit():
        nuevo_username = form.nuevo_username.data

        existing_user = User.query.filter_by(username=nuevo_username, role=current_user.role).first()
        if existing_user:
            flash('Ya existe un usuario con ese nick en el mismo rol.', 'danger')
            return redirect(url_for('edit_username'))

        try:
            old_username = current_user.username
            
            user_profile = UserProfile.query.filter_by(user_id=current_user.id).first()
            if user_profile:
                user_profile.user_name = nuevo_username
            else:
                user_profile = UserProfile(
                    user_name=nuevo_username,
                    user_id=current_user.id
                )
                db.session.add(user_profile)
            
            current_user.username = nuevo_username

            log = UsernameChangeLog(
                changed_by_user_id=current_user.id,
                affected_user_id=current_user.id,
                old_username=old_username,
                new_username=nuevo_username
            )
            db.session.add(log)

            db.session.commit()
            flash('Username actualizado exitosamente.', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error al actualizar: {str(e)}', 'danger')

    return render_template('edit_username.html', form=form)


@app.route('/edit_user_username/<int:user_id>', methods=['GET', 'POST']) 
@login_required
def edit_user_username(user_id):
    if current_user.role != UserRole.ADMIN:
        flash("Acceso denegado. Solo el administrador puede editar usernames.", "danger")
        return redirect(url_for('index'))

    user = User.query.get_or_404(user_id)
    form = EditUsernameForm()

    if form.validate_on_submit():
        nuevo_username = form.nuevo_username.data

        existing_user = User.query.filter_by(username=nuevo_username, role=user.role).first()
        if existing_user:
            flash('Ya existe un usuario con ese nick en el mismo rol.', 'danger')
            return redirect(url_for('edit_user_username', user_id=user.id))

        try:
            old_username = user.username
            
            user_profile = UserProfile.query.filter_by(user_id=user.id).first()
            if user_profile:
                user_profile.user_name = nuevo_username
            else:
                user_profile = UserProfile(
                    user_name=nuevo_username,
                    user_id=user.id
                )
                db.session.add(user_profile)
            
            user.username = nuevo_username

            log = UsernameChangeLog(
                changed_by_user_id=current_user.id,
                affected_user_id=user.id,
                old_username=old_username,
                new_username=nuevo_username
            )
            db.session.add(log)

            db.session.commit()
            flash('Username del usuario actualizada exitosamente.', 'success')
            return redirect(url_for('manage_users'))
        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error al actualizar: {str(e)}', 'danger')

    return render_template('edit_user_username.html', form=form, user=user)


@app.route('/logs_username') 
@login_required
def logs_username():
    if current_user.role != UserRole.ADMIN:
        flash("Acceso denegado. Solo el administrador puede ver los registros de cambios.", "danger")
        return redirect(url_for('index'))

    logs = UsernameChangeLog.query.order_by(UsernameChangeLog.timestamp.desc()).all()
    return render_template('logs.html', logs=logs)


@app.route('/export_logs_excel') 
@login_required
def export_logs_excel():
    if current_user.role != UserRole.ADMIN:
        flash("Acceso denegado. Solo el administrador puede exportar los registros.", "danger")
        return redirect(url_for('index'))

    logs = UsernameChangeLog.query.order_by(UsernameChangeLog.timestamp.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cambios de Username"

    headers = ["#", "Usuario que hizo el cambio", "Usuario afectado", "Username anterior", "Username nuevo", "Fecha y Hora"]
    ws.append(headers)

    for index, log in enumerate(logs, start=1):
        ws.append([
            index,
            log.changed_by.username,
            log.affected_user.username,
            log.old_username,
            log.new_username,
            log.timestamp.strftime('%d/%m/%Y %H:%M:%S')
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="cambios_username.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/export_logs_pdf')
@login_required
def export_logs_pdf():
    """
    Export username change logs to PDF format (Admin only).

    Generates a formatted PDF document with:
    - University logo
    - Table of all username change events
    - Page numbers and timestamps
    - Professional formatting

    Returns:
        PDF File: Download of cambios_username.pdf

    Access Control:
        Only users with ADMIN role can access this route
    """
    if current_user.role != UserRole.ADMIN:
        flash("Acceso denegado. Solo el administrador puede exportar los registros.", "danger")
        return redirect(url_for('index'))

    logs = UsernameChangeLog.query.order_by(UsernameChangeLog.timestamp.desc()).all()
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    logo_path = os.path.join(app.root_path, 'static', 'logo.png')
    if os.path.exists(logo_path):
        c.drawImage(
        logo_path,
        x=2 * cm,
        y=height - 4.5 * cm,
        width=3 * cm,
        height=3 * cm,
        preserveAspectRatio=True,
        mask='auto'
    )

    c.setFont("Helvetica-Bold", 16)
    c.drawString(6 * cm, height - 2.5 * cm, "Historial de Cambios de Username")

    c.setFont("Helvetica-Oblique", 10)
    c.setFillColor(colors.darkgray)
    c.drawString(6 * cm, height - 3.2 * cm, "Sistema de Votación Electrónica - USTA ")
    c.setFillColor(colors.black) 

    c.setFont("Helvetica", 9)
    y = height - 5.2 * cm
    line_height = 1 * cm
    headers = ["#", "Cambió", "Afectado", "Antes", "Ahora", "Fecha y Hora"]
    col_positions = [1.5, 3.5, 7.0, 10.0, 12.5, 15.0]

    def draw_table_header(y_pos):
        c.setFillColor(colors.whitesmoke)
        c.rect(1.5 * cm, y_pos - 0.3 * cm, width - 3 * cm, 0.7 * cm, fill=1)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 9)
        for i, header in enumerate(headers):
            c.drawString(col_positions[i] * cm, y_pos, header)

    def draw_footer(page_num):
        c.setFont("Helvetica-Oblique", 8)
        c.setFillColor(colors.grey)
        c.drawRightString(width - 2 * cm, 1.5 * cm, f"Página {page_num}")
        c.drawString(2 * cm, 1.5 * cm, "Generado por el Sistema de Votación - USTA - Tunja 2025")

    page_num = 1
    draw_table_header(y)
    y -= line_height

    c.setFont("Helvetica", 9)
    for i, log in enumerate(logs, start=1):
        if y < 3 * cm:
            draw_footer(page_num)
            c.showPage()
            page_num += 1
            y = height - 5.2 * cm
            draw_table_header(y)
            y -= line_height

        if i % 2 == 0:
            c.setFillColorRGB(0.95, 0.95, 0.95)
            c.rect(1.5 * cm, y - 0.3 * cm, width - 3 * cm, 0.7 * cm, fill=1)
        c.setFillColor(colors.black)

        row = [
            str(i),
            log.changed_by.username,
            log.affected_user.username,
            log.old_username,
            log.new_username,
            log.timestamp.strftime('%d/%m/%Y %H:%M')
        ]
        for j, text in enumerate(row):
            c.drawString(col_positions[j] * cm, y, text)
        y -= line_height

    draw_footer(page_num)
    c.showPage()
    c.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="cambios_username.pdf",
        mimetype='application/pdf'
    )

@app.route('/logs_password')
@login_required
def logs_password():
    if current_user.role != UserRole.ADMIN:
        flash("Acceso denegado. Solo el administrador puede ver los registros de cambios.", "danger")
        return redirect(url_for('index'))

    logs = PasswordChangeLog.query.order_by(PasswordChangeLog.timestamp.desc()).all()
    return render_template('logs_password.html', logs=logs)

@app.route('/all_logs')
@login_required
def all_logs():
    if current_user.role != UserRole.ADMIN:
        flash("Acceso denegado. Solo el administrador puede ver los registros de cambios.", "danger")
        return redirect(url_for('index'))

    username_logs = UsernameChangeLog.query.order_by(UsernameChangeLog.timestamp.desc()).all()
    password_logs = PasswordChangeLog.query.order_by(PasswordChangeLog.timestamp.desc()).all()
    return render_template('all_logs.html', username_logs=username_logs, password_logs=password_logs)

@app.route('/export_password_logs_excel')
@login_required
def export_password_logs_excel():
    if current_user.role != UserRole.ADMIN:
        flash("Acceso denegado. Solo el administrador puede exportar los registros.", "danger")
        return redirect(url_for('index'))

    logs = PasswordChangeLog.query.order_by(PasswordChangeLog.timestamp.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cambios de Contraseña"

    headers = ["#", "Usuario que hizo el cambio", "Usuario afectado", "Fecha y Hora"]
    ws.append(headers)

    for index, log in enumerate(logs, start=1):
        ws.append([
            index,
            log.changed_by.username,
            log.affected_user.username,
            log.timestamp.strftime('%d/%m/%Y %H:%M:%S')
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="cambios_contraseña.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/export_password_logs_pdf')
@login_required
def export_password_logs_pdf():
    if current_user.role != UserRole.ADMIN:
        flash("Acceso denegado. Solo el administrador puede exportar los registros.", "danger")
        return redirect(url_for('index'))

    logs = PasswordChangeLog.query.order_by(PasswordChangeLog.timestamp.desc()).all()
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    logo_path = os.path.join(app.root_path, 'static', 'logo.png')
    if os.path.exists(logo_path):
        c.drawImage(
            logo_path,
            x=2 * cm,
            y=height - 4.5 * cm,
            width=3 * cm,
            height=3 * cm,
            preserveAspectRatio=True,
            mask='auto'
        )

    c.setFont("Helvetica-Bold", 16)
    c.drawString(6 * cm, height - 2.5 * cm, "Historial de Cambios de Contraseña")

    c.setFont("Helvetica-Oblique", 10)
    c.setFillColor(colors.darkgray)
    c.drawString(6 * cm, height - 3.2 * cm, "Sistema de Votación Electrónica - USTA ")
    c.setFillColor(colors.black)

    c.setFont("Helvetica", 9)
    y = height - 5.2 * cm
    line_height = 1 * cm
    headers = ["#", "Cambió", "Afectado", "Fecha y Hora"]
    col_positions = [1.5, 3.5, 9.0, 15.0]

    def draw_table_header(y_pos):
        c.setFillColor(colors.whitesmoke)
        c.rect(1.5 * cm, y_pos - 0.3 * cm, width - 3 * cm, 0.7 * cm, fill=1)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 9)
        for i, header in enumerate(headers):
            c.drawString(col_positions[i] * cm, y_pos, header)

    def draw_footer(page_num):
        c.setFont("Helvetica-Oblique", 8)
        c.setFillColor(colors.grey)
        c.drawRightString(width - 2 * cm, 1.5 * cm, f"Página {page_num}")
        c.drawString(2 * cm, 1.5 * cm, "Generado por el Sistema de Votación - USTA - Tunja 2025")

    page_num = 1
    draw_table_header(y)
    y -= line_height

    c.setFont("Helvetica", 9)
    for i, log in enumerate(logs, start=1):
        if y < 3 * cm:
            draw_footer(page_num)
            c.showPage()
            page_num += 1
            y = height - 5.2 * cm
            draw_table_header(y)
            y -= line_height

        if i % 2 == 0:
            c.setFillColorRGB(0.95, 0.95, 0.95)
            c.rect(1.5 * cm, y - 0.3 * cm, width - 3 * cm, 0.7 * cm, fill=1)
        c.setFillColor(colors.black)

        row = [
            str(i),
            log.changed_by.username,
            log.affected_user.username,
            log.timestamp.strftime('%d/%m/%Y %H:%M')
        ]
        for j, text in enumerate(row):
            c.drawString(col_positions[j] * cm, y, text)
        y -= line_height

    draw_footer(page_num)
    c.showPage()
    c.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="cambios_contraseña.pdf",
        mimetype='application/pdf'
    )

@app.route('/export_all_logs_excel')
@login_required
def export_all_logs_excel():
    if current_user.role != UserRole.ADMIN:
        flash("Acceso denegado. Solo el administrador puede exportar los registros.", "danger")
        return redirect(url_for('index'))

    username_logs = UsernameChangeLog.query.order_by(UsernameChangeLog.timestamp.desc()).all()
    password_logs = PasswordChangeLog.query.order_by(PasswordChangeLog.timestamp.desc()).all()

    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "Cambios de Username"
    headers1 = ["#", "Usuario que hizo el cambio", "Usuario afectado", "Username anterior", "Username nuevo", "Fecha y Hora"]
    ws1.append(headers1)

    for index, log in enumerate(username_logs, start=1):
        ws1.append([
            index,
            log.changed_by.username,
            log.affected_user.username,
            log.old_username,
            log.new_username,
            log.timestamp.strftime('%d/%m/%Y %H:%M:%S')
        ])

    ws2 = wb.create_sheet(title="Cambios de Contraseña")
    headers2 = ["#", "Usuario que hizo el cambio", "Usuario afectado", "Fecha y Hora"]
    ws2.append(headers2)

    for index, log in enumerate(password_logs, start=1):
        ws2.append([
            index,
            log.changed_by.username,
            log.affected_user.username,
            log.timestamp.strftime('%d/%m/%Y %H:%M:%S')
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="logs_sistema.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/export_all_logs_pdf')
@login_required
def export_all_logs_pdf():
    if current_user.role != UserRole.ADMIN:
        flash("Acceso denegado. Solo el administrador puede exportar los registros.", "danger")
        return redirect(url_for('index'))

    username_logs = UsernameChangeLog.query.order_by(UsernameChangeLog.timestamp.desc()).all()
    password_logs = PasswordChangeLog.query.order_by(PasswordChangeLog.timestamp.desc()).all()
    
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    def draw_footer(page_num):
        c.setFont("Helvetica-Oblique", 8)
        c.setFillColor(colors.grey)
        c.drawRightString(width - 2 * cm, 1.5 * cm, f"Página {page_num}")
        c.drawString(2 * cm, 1.5 * cm, "Generado por el Sistema de Votación - USTA - Tunja 2025")

    def draw_header(title):
        logo_path = os.path.join(app.root_path, 'static', 'logo.png')
        if os.path.exists(logo_path):
            c.drawImage(
                logo_path,
                x=2 * cm,
                y=height - 4.5 * cm,
                width=3 * cm,
                height=3 * cm,
                preserveAspectRatio=True,
                mask='auto'
            )

        c.setFont("Helvetica-Bold", 16)
        c.drawString(6 * cm, height - 2.5 * cm, title)

        c.setFont("Helvetica-Oblique", 10)
        c.setFillColor(colors.darkgray)
        c.drawString(6 * cm, height - 3.2 * cm, "Sistema de Votación Electrónica - USTA ")
        c.setFillColor(colors.black)

    page_num = 1
    draw_header("Historial de Cambios de Username")
    
    c.setFont("Helvetica", 9)
    y = height - 5.2 * cm
    line_height = 1 * cm
    
    headers1 = ["#", "Cambió", "Afectado", "Antes", "Ahora", "Fecha y Hora"]
    col_positions1 = [1.5, 3.5, 7.0, 10.0, 12.5, 15.0]

    def draw_username_header(y_pos):
        c.setFillColor(colors.whitesmoke)
        c.rect(1.5 * cm, y_pos - 0.3 * cm, width - 3 * cm, 0.7 * cm, fill=1)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 9)
        for i, header in enumerate(headers1):
            c.drawString(col_positions1[i] * cm, y_pos, header)

    draw_username_header(y)
    y -= line_height

    c.setFont("Helvetica", 9)
    for i, log in enumerate(username_logs, start=1):
        if y < 3 * cm:
            draw_footer(page_num)
            c.showPage()
            page_num += 1
            draw_header("Historial de Cambios de Username (Continuación)")
            y = height - 5.2 * cm
            draw_username_header(y)
            y -= line_height

        if i % 2 == 0:
            c.setFillColorRGB(0.95, 0.95, 0.95)
            c.rect(1.5 * cm, y - 0.3 * cm, width - 3 * cm, 0.7 * cm, fill=1)
        c.setFillColor(colors.black)

        row = [
            str(i),
            log.changed_by.username,
            log.affected_user.username,
            log.old_username,
            log.new_username,
            log.timestamp.strftime('%d/%m/%Y %H:%M')
        ]
        for j, text in enumerate(row):
            c.drawString(col_positions1[j] * cm, y, text)
        y -= line_height

    draw_footer(page_num)
    c.showPage()
    
    page_num += 1
    draw_header("Historial de Cambios de Contraseña")
    
    headers2 = ["#", "Cambió", "Afectado", "Fecha y Hora"]
    col_positions2 = [1.5, 3.5, 9.0, 15.0]
    
    y = height - 5.2 * cm
    
    def draw_password_header(y_pos):
        c.setFillColor(colors.whitesmoke)
        c.rect(1.5 * cm, y_pos - 0.3 * cm, width - 3 * cm, 0.7 * cm, fill=1)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 9)
        for i, header in enumerate(headers2):
            c.drawString(col_positions2[i] * cm, y_pos, header)

    draw_password_header(y)
    y -= line_height

    c.setFont("Helvetica", 9)
    for i, log in enumerate(password_logs, start=1):
        if y < 3 * cm:
            draw_footer(page_num)
            c.showPage()
            page_num += 1
            draw_header("Historial de Cambios de Contraseña (Continuación)")
            y = height - 5.2 * cm
            draw_password_header(y)
            y -= line_height

        if i % 2 == 0:
            c.setFillColorRGB(0.95, 0.95, 0.95)
            c.rect(1.5 * cm, y - 0.3 * cm, width - 3 * cm, 0.7 * cm, fill=1)
        c.setFillColor(colors.black)

        row = [
            str(i),
            log.changed_by.username,
            log.affected_user.username,
            log.timestamp.strftime('%d/%m/%Y %H:%M')
        ]
        for j, text in enumerate(row):
            c.drawString(col_positions2[j] * cm, y, text)
        y -= line_height

    draw_footer(page_num)
    c.showPage()
    c.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="todos_los_logs.pdf",
        mimetype='application/pdf'
    )

if __name__ == '__main__':
    app.run(debug=True)
